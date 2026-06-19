"""
NSE Stock Data Fetcher using NseIndiaApi
Handles NSE's session requirements automatically
"""

from datetime import datetime
import os
import sqlite3
from pathlib import Path
from typing import Any

from nse import NSE
from openpyxl import Workbook, load_workbook

stocks = [

"RELIANCE",
"HDFCBANK",
"TCS",
"BHARTIARTL",
"ICICIBANK",
"SBIN",
"INFY",
"LICI",
"BAJFINANCE",
"HINDUNILVR",
"ITC",
"LT",
"HCLTECH",
"KOTAKBANK",
"MARUTI",
"SUNPHARMA",
"AXISBANK",
"ULTRACEMCO",
"HAL",
"NTPC",
"BAJAJFINSV",
"ONGC",
"TITAN",
"ADANIPORTS",
"BEL",
"ADANIENT",
"DMART",
"WIPRO",
"POWERGRID",
"TMCV",
"JSWSTEEL",
"COALINDIA",
"BAJAJ-AUTO",
"NESTLEIND",
"ASIANPAINT",
"ADANIPOWER",
"DLF",
"INDIGO",
"TRENT",
"IOC",
"HINDZINC",
"TATASTEEL",
"JIOFIN",
"GRASIM",
"SBILIFE",
"VEDL",
"IRFC",
"DIVISLAB",
"TECHM",
"HDFCLIFE",
"LTIM",
"VBL",
"HYUNDAI",
"SOLARINDS",
"ADANIGREEN",
"BAJAJHLDNG",
"PIDILITIND",
"EICHERMOT",
"HINDALCO",
"LODHA",
"BPCL",
"BRITANNIA",
"AMBUJACEM",
"CHOLAFIN",
"TVSMOTOR",
"MAZDOCK",
"PFC",
"ABB",
"TATAPOWER",
"SHRIRAMFIN",
"GODREJCP",
"BANKBARODA",
"GAIL",
"PNB",
"CIPLA",
"SIEMENS",
"MAXHEALTH",
"UNIONBANK",
"DRREDDY",
"TORNTPHARM",
"INDHOTEL",
"BSE",
"UNITDSPR",
"HDFCAMC",
"SHREECEM",
"MUTHOOTFIN",
"TATACONSUM",
"MOTHERSON",
"CGPOWER",
"RECLTD",
"INDUSTOWER",
"ADANIENSOL",
"BAJAJHFL",
"APOLLOHOSP",
"IDBI",
"CANBK",
"MANKIND",
"ZYDUSLIFE",
"HAVELLS",
"ICICIGI",
"BOSCHLTD",
"NAUKRI",
"PERSISTENT",
"SBICARD",
"ICICIPRULI",
"JINDALSTEL",
"SWIGGY",
"SRF",
"CUMMINSIND",
"POLYCAB",
"NTPCGREEN",
"MARICO",
"LUPIN",
"JSWENERGY",
"BHEL",
"SUZLON",
"GMRAIRPORT",
"HEROMOTOCO",
"BHARTIHEXA",
"POLICYBZR",
"DIXON",
"INDIANB",
"HINDPETRO",
"OFSS",
"DABUR",
"NHPC",
"RVNL",
"POWERINDIA",
"LLOYDSME",
"WAAREEENER",
"OIL",
"GODREJPROP",
"PRESTIGE",
"IOB",
"ATGL",
"IDEA",
"TORNTPOWER",
"BDL",
"OBEROIRLTY",
"ASHOKLEY",
"COROMANDEL",
"ABCAPITAL",
"FACT",
"INDUSINDBK",
"COLPAL",
"GICRE",
"BERGEPAINT",
"JSWINFRA",
"AUROPHARMA",
"PIIND",
"YESBANK",
"SCHAEFFLER",
"BHARATFORG",
"COFORGE",
"IRCTC",
"UNOMINDA",
"PATANJALI",
"NMDC",
"LINDEINDIA",
"AUBANK",
"MRF",
"SUPREMEIND",
"ALKEM",
"PHOENIXLTD",
"FORTIS",
"VMM",
"PAYTM",
"COCHINSHIP",
"TIINDIA",
"NYKAA",
"GLAXO",
"JSL",
"MFSL",
"BANKINDIA",
"KALYANKJIL",
"SUNDARMFIN",
"SAIL",
"UBL",
"IDFCFIRSTB",
"MPHASIS",
"FEDERALBNK",
"PAGEIND",
"APLAPOLLO",
"MOTILALOFS",
"TATACOMM",
"UPL",
"BALKRISIND",
"LTF",
"NAM-INDIA",
"GLENMARK",
"LTTS",
"IREDA",
"CONCOR",
"GODREJIND",
"JKCEMENT",
"JUBLFOOD",
"PREMIERENE",
"HUDCO",
"PETRONET",
"360ONE",
"PGHH",
"AIIL",
"VOLTAS",
"GODFRYPHLP",
"THERMAX",
"ASTRAL",
"MAHABANK",
"CRISIL",
"BIOCON",
"TATAELXSI",
"KPITTECH",
"UCOBANK",
"NH",
"FLUOROCHEM",
"DALBHARAT",
"KPRMILL",
"CHOLAHLDNG",
"SJVN",
"M&MFIN",
"GRSE",
"KAYNES",
"ESCORTS",
"LAURUSLABS",
"RADICO",
"CDSL",
"ACC",
"IPCALAB",
"CENTRALBK",
"AWL",
"KEI",
"NATIONALUM",
"ENDURANCE",
"GILLETTE",
"BLUESTARCO",
"HONAUT",
"LICHSGFIN",
"TATAINVEST",
"3MINDIA",
"POONAWALLA",
"AJANTPHARM",
"METROBRAND",
"EXIDEIND",
"APARINDS",
"NBCC",
"GUJGASLTD",
"GODIGIT",
"AIAENG",
"MEDANTA",
"NLCINDIA",
"ITI",
"NIACL",
"SONACOMS",
"TATATECH",
"ASTERDM",
"IRB",
"IKS",
"BANDHANBNK",
"BRIGADE",
"IGL",
"GLAND",
"APOLLOTYRE",
"WOCKPHARMA",
"RPOWER",
"PNBHOUSING",
"ABREL",
"DELHIVERY",
"AEGISLOG",
"JBCHEPHARM",
"MSUMI",
"FSL",
"AFFLE",
"KIMS",
"PFIZER",
"NUVAMA",
"DEEPAKNTR",
"SYNGENE",
"ANGELONE",
"PPLPHARMA",
"STARHEALTH",
"JYOTICNC",
"EMCURE",
"PEL",
"ZFCVINDIA",
"SUMICHEM",
"EMAMILTD",
"HINDCOPPER",
"LALPATHLAB",
"SUNTV",
"RAMCOCEM",
"TIMKEN",
"ASTRAZEN",
"ERIS",
"WELCORP",
"JSWHL",
"KEC",
"MRPL",
"TATACHEM",
"NAVINFLUOR",
"REDINGTON",
"SHYAMMETL",
"SKFINDIA",
"AMBER",
"MANAPPURAM",
"FIVESTAR",
"CROMPTON",
"HSCL",
"EIHOTEL",
"PTCIL",
"INOXWIND",
"CHAMBLFERT",
"PSB",
"BASF",
"CESC",
"ABSLAMC",
"POLYMED",
"PGEL",
"CONCORDBIO",
"HATSUN",
"TVSHLTD",
"KFINTECH",
"ATUL",
"DEVYANI",
"SUNDRMFAST",
"CASTROLIND",
"IIFL",
"KPIL",
"CAMS",
"OLAELEC",
"VINATIORGA",
"RATNAMANI",
"KARURVYSYA",
"CHALET",
"MANYAVAR",
"KANSAINER",
"ZENSARTECH",
"SAGILITY",
"TRITURBINE",
"DEEPAKFERT",
"FIRSTCRY",
"AADHARHFC",
"APLLTD",
"ASAHIINDIA",
"GRINDWELL",
"SCHNEIDER",
"BEML",
"IRCON",
"GSPL",
"CREDITACC",
"DCMSHRIRAM",
"FORCEMOT",
"ANANTRAJ",
"JUBLPHARMA",
"BIKAJI",
"SIGNATURE",
"CARBORUNIV",
"KIOCL",
"ZENTEC",
"TECHNOE",
"ECLERX",
"LMW",
"WHIRLPOOL",
"ANANDRATHI",
"EIDPARRY",
"NEWGEN",
"CIEINDIA",
"DATAPATTNS",
"CENTURYPLY",
"SOBHA",
"INTELLECT",
"ELGIEQUIP",
"VGUARD",
"JBMA",
"IEX",
"VENTIVE",
"MAHSCOOTER",
"AARTIIND",
"CGCL",
"KAJARIACER",
"AFCONS",
"IFCI",
"JWL",
"UTIAMC",
"HBLENGINE",
"SARDAEN",
"CAPLIPOINT",
"APTUS",
"NEULANDLAB",
"NATCOPHARM",
"NAVA",
"IGIL",
"BATAINDIA",
"EMBDL",
"ACMESOLAR",
"SAILIFE",
"CLEAN",
"TRIDENT",
"BLUEJET",
"LTFOODS",
"JINDALSAW",
"CYIENT",
"SPLPETRO",
"RRKABEL",
"BLS",
"NIVABUPA",
"DOMS",
"INDIAMART",
"RAINBOW",
"KSB",
"BLUEDART",
"FINCABLES",
"GODREJAGRO",
"PCBL",
"TARIL",
"CEATLTD",
"AAVAS",
"FINPIPE",
"ELECON",
"CUB",
"KIRLOSBROS",
"ACE",
"FINEORG",
"INDGN",
"NCC",
"MGL",
"SANOFI",
"RBLBANK",
"CHOICEIN",
"GESHIP",
"TBOTEK",
"RAILTEL",
"JMFINANCIL",
"PARADEEP",
"VTL",
"RITES",
"BBTC",
"CELLO",
"HOMEFIRST",
"TTML",
"ZYDUSWELL",
"MINDACORP",
"ZEEL",
"WELSPUNLIV",
"VOLTAMP",
"GMDCLTD",
"GRINFRA",
"CRAFTSMAN",
"ENGINERSIN",
"KIRLOSENG",
"GRAVITA",
"ANURAS",
"NUVOCO",
"TEJASNET",
"JYOTHYLAB",
"INGERRAND",
"GPIL",
"SHAKTIPUMP",
"GRANULES",
"ABDL",
"DBREALTY",
"BALRAMCHIN",
"HFCL",
"JPPOWER",
"BSOFT",
"TITAGARH",
"SANOFICONR",
"SBFC",
"SONATSOFTW",
"NAZARA",
"RKFORGE",
"EUREKAFORB",
"JUBLINGREA",
"MARKSANS",
"PRUDENT",
"NSLNISP",
"GALLANTT",
"VESUVIUS",
"WESTLIFE",
"GENUSPOWER",
"GRAPHITE",
"LEMONTREE",
"ASTRAMICRO",
"CCL",
"INOXINDIA",
"AZAD",
"SHRIPISTON",
"CANFINHOME",
"MEDPLUS",
"SAFARI",
"RELAXO",
"RHIM",
"SCI",
"BIRLACORPN",
"SAMMAANCAP",
"AETHER",
"SAREGAMA",
"EDELWEISS",
"INDIACEM",
"JKTYRE"
]


def get_value(data: dict[str, Any], path: tuple[str, ...], default: Any = None) -> Any:
    current = data
    for key in path:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
        if current in (None, ""):
            return default
    return current


def first_value(*values: Any, default: Any = None) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return default


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def fetch_quote(nse: NSE, symbol: str) -> dict[str, Any]:
    if hasattr(nse, "quote"):
        return nse.quote(symbol)
    if hasattr(nse, "equityQuote"):
        return nse.equityQuote(symbol)
    if hasattr(nse, "equity_quote"):
        return nse.equity_quote(symbol)
    raise AttributeError("NSE object does not have quote(), equityQuote(), or equity_quote()")


def get_stock_data(nse: NSE, symbol: str) -> dict[str, Any] | None:
    try:
        print(f"Fetching {symbol} data from NSE...")
        quote = fetch_quote(nse, symbol)

        if not quote:
            print(f"NSE returned an empty response for {symbol}.")
            return None

        meta = quote.get("metaData", {})
        trade = quote.get("tradeInfo", {})
        sec = quote.get("secInfo", {})

        last_price = to_float(
            first_value(
                trade.get("lastPrice"),
                quote.get("lastPrice"),
                meta.get("closePrice"),
                meta.get("lastPrice"),
                default=0,
            )
        )

        change = to_float(meta.get("change", 0))
        percentage_change = to_float(meta.get("pChange", 0))
        market_cap = to_float(trade.get("totalMarketCap", 0))
        volume_traded = to_int(first_value(trade.get("totalTradedVolume"), trade.get("quantitytraded")))
        delivery_quantity = to_int(first_value(sec.get("deliveryQuantity"), trade.get("deliveryquantity")))
        delivery_percentage = to_float(
            first_value(
                trade.get("deliveryToTradedQuantity"),
                sec.get("deliveryTotradedQuantity"),
                default=0,
            )
        )

        return {
            "symbol": symbol,
            "company_name": meta.get("companyName", symbol),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "last_price": last_price,
            "change": change,
            "percentage_change": percentage_change,
            "market_cap": market_cap,
            "market_cap_cr": round(market_cap / 10_000_000, 2) if market_cap else 0,
            "volume_traded": volume_traded,
            "delivery_quantity": delivery_quantity,
            "delivery_percentage": delivery_percentage,
            "open": to_float(meta.get("open", 0)),
            "high": to_float(meta.get("dayHigh", 0)),
            "low": to_float(meta.get("dayLow", 0)),
            "previous_close": to_float(meta.get("previousClose", 0)),
            "last_update_time": quote.get("lastUpdateTime"),
        }

    except Exception as e:
        print(f"Error fetching {symbol}: {e}")
        return None


def display_stock(data: dict[str, Any]):
    print("\n" + "=" * 70)
    print(f"{data['company_name']} ({data['symbol']})")
    print("=" * 70)
    print(f"As on: {data['timestamp']} IST")
    print("-" * 70)
    print(f"Last Price: Rs {data['last_price']:,.2f}")
    print(f"Change: {data['change']:+,.2f} ({data['percentage_change']:+.2f}%)")
    print(f"Open: Rs {data['open']:,.2f}")
    print(f"High: Rs {data['high']:,.2f}")
    print(f"Low: Rs {data['low']:,.2f}")
    print(f"Previous Close: Rs {data['previous_close']:,.2f}")
    print(f"Market Cap: Rs {data['market_cap_cr']:,.2f} Cr")
    print(f"Delivery %: {data['delivery_percentage']:.2f}%")
    print(f"Volume: {data['volume_traded']:,} shares")
    print(f"Last Update: {data.get('last_update_time') or data['timestamp']}")
    print("=" * 70)


def save_to_excel(sorted_results, date):
    stock_row = {}

    if os.path.exists('Stcapmark.xlsx'):
        wb = load_workbook('Stcapmark.xlsx')
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            stock = ws.cell(row=row, column=1).value
            if stock:
                stock_row[stock] = row
        market_col = None
        col = 2
        while ws.cell(row=1, column=col).value is not None:
            if ws.cell(row=1, column=col).value == f'{date} Market Cap':
                market_col = col
                delivery_col = col + 1
                rank_col = col + 2
                break
            col += 3
        if market_col is None:
            col = 2
            while ws.cell(row=1, column=col).value is not None:
                col += 1
            market_col = col
            delivery_col = col + 1
            rank_col = col + 2
            ws.cell(row=1, column=market_col, value=f'{date} Market Cap')
            ws.cell(row=1, column=delivery_col, value=f'{date} Delivery Qty')
            ws.cell(row=1, column=rank_col, value=f'{date} Rank')
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='Stock Name')
        for row, stock in enumerate(stocks, 2):
            ws.cell(row=row, column=1, value=stock)
            stock_row[stock] = row
        market_col = 2
        delivery_col = 3
        rank_col = 4
        ws.cell(row=1, column=market_col, value=f'{date} Market Cap')
        ws.cell(row=1, column=delivery_col, value=f'{date} Delivery Qty')
        ws.cell(row=1, column=rank_col, value=f'{date} Rank')

    for item in sorted_results:
        row = stock_row.get(item['stock'])
        if row:
            if 'element_int' in item:
                ws.cell(row=row, column=market_col, value=item['element_int'])
            if 'delivery' in item:
                ws.cell(row=row, column=delivery_col, value=item['delivery'])
            ws.cell(row=row, column=rank_col, value=item['rank'])

    wb.save('Stcapmark.xlsx')
    print(f"Data saved to Stcapmark.xlsx ({len(sorted_results)} stocks)")


def save_to_database(sorted_results, date):
    db_path = 'stock_data.db'
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stock_market_cap (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stock TEXT NOT NULL,
            market_cap INTEGER,
            delivery_percentage REAL,
            rank INTEGER,
            date TEXT NOT NULL
        )
    ''')
    
    for item in sorted_results:
        cursor.execute('''
            INSERT OR REPLACE INTO stock_market_cap (stock, market_cap, delivery_percentage, rank, date)
            VALUES (?, ?, ?, ?, ?)
        ''', (item['stock'], item.get('element_int'), item.get('delivery'), item.get('rank'), date))
    
    conn.commit()
    conn.close()
    print(f"Data saved to {db_path}")


def main():
    print("\nNSE Stock Data Fetcher - Multiple Stocks")

    results = []
    with NSE(download_folder=Path.cwd(), server=False, timeout=20) as nse:
        for symbol in stocks:
            data = get_stock_data(nse, symbol)
            if data:
                element_int = int(data['market_cap'] / 10_000_000) if data['market_cap'] else 0
                results.append({'stock': symbol, 'element_int': element_int, 'delivery': data['delivery_percentage']})
                print(f"{symbol}: Market Cap = Rs {data['market_cap_cr']:,.2f} Cr, Delivery % = {data['delivery_percentage']:.2f}%")

    # Sort and rank by market cap descending
    sorted_results = sorted(results, key=lambda x: x.get('element_int', 0) if x.get('element_int') else 0, reverse=True)
    for i, item in enumerate(sorted_results, 1):
        item['rank'] = i

    date = datetime.now().date().isoformat()
    
    # Save to Excel
    save_to_excel(sorted_results, date)
    
    # Save to SQLite database
    save_to_database(sorted_results, date)


if __name__ == "__main__":
    main()